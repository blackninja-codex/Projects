import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook

#flask can be used with the frontend and the linkedIn user and recipient users credential can be appended in the list with it to incorporate new users, rather than hard coding number of users as done in this case.

# 2 users are used for testing purpose

linkedin_users = [
    {"email": "user1@gmail.com", "password": bcrypt.hashpw(b"password1", bcrypt.gensalt())},
    {"email": "user2@gmail.com", "password": bcrypt.hashpw(b"password2", bcrypt.gensalt())}
]
recipient_emails = [
    "recipient1@gmail.com",
    "recipient2@gmail.com"
]
excel_file_path = "linkedin_data.xlsx"
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")

def login_to_linkedin(email, password):
    driver = webdriver.Chrome(options=chrome_options)
    driver.get("https://www.linkedin.com")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "session_key")))
    driver.find_element(By.ID, "session_key").send_keys(email)
    stored_password = next((user["password"] for user in linkedin_users if user["email"] == email), None)
    if stored_password and bcrypt.checkpw(password.encode(), stored_password):
        driver.find_element(By.ID, "session_password").send_keys(password)
        driver.find_element(By.CLASS_NAME, "sign-in-form__submit-button").click()
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "msg-overlay-bubble-header")))
    else:
        print("Invalid email or password")
        driver.quit()
    return driver

def extract_linkedin_data(driver):
    driver.get("https://www.linkedin.com/in/your-profile")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "pv-profile-section")))
    messages_count = driver.find_element(By.CLASS_NAME, "msg-overlay-bubble-header").text
    notifications_count = driver.find_element(By.CLASS_NAME, "global-nav__inbox-icon").text
    return messages_count, notifications_count

def send_email_notification(recipient_email, messages_count, notifications_count, comparison_data):
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "your-email@gmail.com"
    sender_password = "your-password"
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = recipient_email
    message["Subject"] = "LinkedIn Notifications"
    email_body = f"Number of unread messages: {messages_count}\n"
    email_body += f"Number of unread notifications: {notifications_count}\n"
    email_body += "Comparison with previous occurrence:\n"
    email_body += comparison_data
    message.attach(MIMEText(email_body, "plain"))
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)

def compare_data(current_data, previous_data):
    new_messages_count = current_data[0] - previous_data[0]
    new_notifications_count = current_data[1] - previous_data[1]
    comparison_result = f"New messages received in the last 3 hours: {new_messages_count}\n"
    comparison_result += "Comparison result"
    return comparison_result

def main():
    for user in linkedin_users:
        driver = login_to_linkedin(user["email"], user["password"])
        messages_count, notifications_count = extract_linkedin_data(driver)
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active
        previous_data = worksheet.cell(row=2, column=1).value
        comparison_result = compare_data((messages_count, notifications_count), previous_data)
        worksheet.cell(row=2, column=1).value = (messages_count, notifications_count)
        workbook.save(excel_file_path)
        driver.quit()
        for recipient_email in recipient_emails:
            send_email_notification(recipient_email, messages_count, notifications_count, comparison_result)

if __name__ == "__main__":
    main()
    
    while True:
        time.sleep(3 * 60 * 60)
        if __name__ == "__main__":
            main()
